#!/usr/bin/env python3
"""Bundle per-table TSVs in results/ into a single Supplementary_Tables.xlsx.

One sheet per supplementary table, plus a TOC sheet. Run:
    python3 scripts/14_build_supplementary_tables.py
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

PROJECT = Path("/Users/jezmarston/Desktop/Bulk_Acral_Claude")
RESULTS = PROJECT / "results"
OUT     = PROJECT / "docs" / "Supplementary_Tables.xlsx"

# (sheet_name, title, [(sub_label, file)])  -- multi-file tables get sub-headers
TABLES = [
    ("S1_Cohort", "Table S1. Cohort and clinical annotations", [
        ("Cohort summary",        "cohort_summary.tsv"),
        ("Clinical master table", "clinical_master.tsv"),
    ]),
    ("S2_Clusters", "Table S2. Transcriptomic cluster assignments and survival", [
        ("Cluster assignments",          "cluster_assignments.tsv"),
        ("Cluster survival HRs",         "cluster_survival_HRs.tsv"),
        ("Manual cluster survival HRs",  "manual_cluster_survival_HRs.tsv"),
        ("Cluster vs confounders",       "cluster_vs_confounds.tsv"),
        ("k selection summary",          "k_selection_summary.tsv"),
        ("Silhouette by k",              "silhouette_by_k.tsv"),
        ("PAC by k",                     "pac_by_k.tsv"),
    ]),
    ("S3_Cluster_DE_GO", "Table S3. Cluster-specific differential expression and GO enrichment", [
        ("C1 DE",  "cluster_C1_DE.tsv"),
        ("C1 GO",  "cluster_C1_GO.tsv"),
        ("C2 DE",  "cluster_C2_DE.tsv"),
        ("C2 GO",  "cluster_C2_GO.tsv"),
        ("C3 DE",  "cluster_C3_DE.tsv"),
        ("C3 GO",  "cluster_C3_GO.tsv"),
        ("C4 DE",  "cluster_C4_DE.tsv"),
        ("C4 GO",  "cluster_C4_GO.tsv"),
    ]),
    ("S4_Cox", "Table S4. Survival modelling (univariate and multivariable Cox)", [
        ("Univariate Cox - HERV loci", "univariate_cox_HERV.tsv"),
        ("Multivariable Cox",          "multivariable_cox.tsv"),
        ("Time-dependent AUC",         "time_dependent_auc.tsv"),
        ("ICI response logistic",      "ici_response_logistic.tsv"),
    ]),
    ("S5_Signature_Stability", "Table S5. Prognostic HERV signature - feature selection and stability", [
        ("LASSO stability selection",       "lasso_stability_selection.tsv"),
        ("Boruta decisions",                "boruta_decisions.tsv"),
        ("Bootstrap Jaccard stability",     "bootstrap_jaccard.tsv"),
        ("CORE vs CORE+HML c-index",        "core_vs_core_plus_hml_cindex.tsv"),
        ("Candidate signature c-index",     "candidate_signature_cindex.tsv"),
    ]),
    ("S6_HLA_typing", "Table S6. HLA class I typing and population frequencies", [
        ("Per-sample typings (long)", "hla_typings_long.tsv"),
        ("Cohort allele frequencies", "hla_allele_frequencies.tsv"),
    ]),
    ("S7_HML6_binders", "Table S7. HML6_20p11.21 - per-patient peptide-HLA binder coverage", [
        ("Per-patient coverage", "hml6_binder_coverage_per_patient.tsv"),
    ]),
    ("S8_HERVH_binders", "Table S8. HERVH_6q21a - per-patient binder coverage and ESRG alignment", [
        ("Per-patient coverage",          "hervh_binder_coverage_per_patient.tsv"),
        ("HERVH-ESRG alignment summary",  "hervh_esrg_alignment_summary.tsv"),
    ]),
    ("S9_Immune_correlations", "Table S9. HERV expression vs immune infiltrate correlations", [
        ("HERV-immune correlations", "herv_immune_correlations.tsv"),
        ("HML-immune correlations",  "hml_immune_correlations.tsv"),
        ("MCP-counter deconvolution","immune_deconvolution_mcp.tsv"),
    ]),
    ("S10_Coding_capacity", "Table S10. Coding capacity and ORF annotation of signature loci", [
        ("Signature ORFs",                "signature_loci_ORFs.tsv"),
        ("Signature locus coordinates",   "signature_locus_coordinates.tsv"),
        ("HML candidate coordinates",     "hml_candidate_coordinates.tsv"),
        ("HML coding summary",            "hml_coding_summary.tsv"),
        ("Signature coding summary",      "signature_coding_summary.tsv"),
        ("Axis members per locus",        "axis_members_per_locus.tsv"),
        ("Oncofetal axis vs CORE",        "oncofetal_axis_vs_core.tsv"),
    ]),
]


def read_tsv(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    return pd.read_csv(path, sep="\t")


def write_block(ws, df: pd.DataFrame, start_row: int, sub_label: str | None) -> int:
    """Write df into ws starting at start_row. Returns next free row."""
    row = start_row
    if sub_label:
        ws.cell(row=row, column=1, value=sub_label).font = Font(bold=True, size=11)
        row += 1
    # header
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=row, column=j, value=str(col))
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    row += 1
    # body
    for _, rec in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            v = rec[col]
            if pd.isna(v):
                continue
            ws.cell(row=row, column=j, value=v if not hasattr(v, "item") else v.item())
        row += 1
    return row + 2  # blank-row separator


def autosize(ws, max_width: int = 60):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None),
                      default=10)
        ws.column_dimensions[letter].width = min(longest + 2, max_width)


def main() -> int:
    if not RESULTS.is_dir():
        print(f"ERROR: {RESULTS} not found", file=sys.stderr); return 1

    # Build TOC entries as we go
    toc_rows = []
    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        # Placeholder so we can re-order TOC to be first
        pd.DataFrame({"_": [""]}).to_excel(xw, sheet_name="TOC", index=False)

        for sheet, title, parts in TABLES:
            # Resolve files; skip missing
            present = [(lbl, RESULTS / fn) for lbl, fn in parts if (RESULTS / fn).exists()]
            missing = [fn for lbl, fn in parts if not (RESULTS / fn).exists()]
            if not present:
                print(f"  [skip] {sheet}: no source files present")
                continue

            # Write a single empty sheet first, then fill via openpyxl directly
            pd.DataFrame().to_excel(xw, sheet_name=sheet, index=False)
            ws = xw.book[sheet]
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
            row = 3
            for lbl, fp in present:
                df = read_tsv(fp)
                sub = f"{lbl}  -  {fp.name}  ({len(df)} rows)"
                row = write_block(ws, df, row, sub)
            autosize(ws)
            toc_rows.append({
                "Sheet":       sheet,
                "Title":       title,
                "Sub-tables":  len(present),
                "Files":       "; ".join(fp.name for _, fp in present),
                "Missing":     "; ".join(missing) if missing else "",
            })
            print(f"  [ok]   {sheet}: {len(present)} sub-table(s)"
                  + (f"  (missing: {missing})" if missing else ""))

        # Now fill the TOC
        toc = xw.book["TOC"]
        toc.delete_rows(1, toc.max_row)
        toc.cell(row=1, column=1,
                 value="Supplementary Tables - Acral melanoma HERV/TE manuscript"
                ).font = Font(bold=True, size=14)
        toc.cell(row=2, column=1,
                 value="Each sheet bundles related per-table TSVs from the Telescope/edgeR/Cox pipeline."
                ).alignment = Alignment(wrap_text=True)
        hdr = ["Sheet", "Title", "Sub-tables", "Files", "Missing"]
        for j, h in enumerate(hdr, start=1):
            c = toc.cell(row=4, column=j, value=h)
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9E1F2")
        for i, rec in enumerate(toc_rows, start=5):
            for j, h in enumerate(hdr, start=1):
                toc.cell(row=i, column=j, value=rec[h])
            # hyperlink Sheet col -> sheet
            link = toc.cell(row=i, column=1)
            link.hyperlink = f"#'{rec['Sheet']}'!A1"
            link.font = Font(color="0563C1", underline="single")
        autosize(toc, max_width=80)

    print(f"\nWrote: {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
